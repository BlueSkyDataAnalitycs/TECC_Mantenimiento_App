#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_taller.py — Importa informes históricos del taller propio (Excel) a taller_seed.json.

Estos informes son la BASE DE DATOS histórica del taller (ene–jun 2026). Quedan
embebidos en la app y SIEMPRE están presentes; las intervenciones nuevas llegan
aparte por el Google Sheet del taller y se unen a estas en tiempo real.

Uso:
    python3 import_taller.py "INFORME ... (1).xlsx" "OTRO INFORME.xlsx" ...
    (sin argumentos usa las rutas por defecto en ~/Downloads)

Formato esperado por hoja (los informes traen secciones por mes):
    · Con fecha:  FECHA | N° BUSETA | CC CONDUCTOR | KILOMETRAJE | INTERVECION
    · Sin fecha:  N° BUSETA | CC CONDUCTOR | KILOMETRAJE | INTERVECION   (mes = encabezado)
Los encabezados de mes ("ABRIL 2026", "FEBRERO 2026"…) pueden aparecer en cualquier columna.
"""
import sys, os, re, json, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS = os.path.expanduser("~/Downloads")
DEFAULT_FILES = [
    os.path.join(DOWNLOADS, "INFORME  DEL  ENE- FEB Y  MARZO (1).xlsx"),
    os.path.join(DOWNLOADS, "INFORME  TCC 2026 (1).xlsx"),
]

MESES = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
         "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12}


def txt(c):
    return ("" if c is None else str(c)).strip()


def mes_de(cells):
    """Si alguna celda empieza con un nombre de mes, devuelve (año, mes); si no, None."""
    for c in cells:
        s = txt(c).upper()
        for nombre, mn in MESES.items():
            if s.startswith(nombre):
                y = re.search(r"(20\d{2})", s)
                return (int(y.group(1)) if y else 2026, mn)
    return None


def es_encabezado_tabla(cells):
    j = " ".join(txt(c).upper() for c in cells)
    return "BUSETA" in j or "INTERVEC" in j


def norm_buseta(v):
    return re.sub(r"\s+", "", txt(v).upper())


def parse_file(fn, records, seq):
    wb = openpyxl.load_workbook(fn, data_only=True)
    ws = wb.active
    # ¿la hoja trae columna FECHA? (5 columnas con FECHA en el encabezado)
    has_date = False
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if any("FECHA" in txt(c).upper() for c in row):
            has_date = True
            break
    cur_y, cur_m = 2026, None
    added = 0
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not any(txt(c) for c in cells):
            continue
        buseta = txt(cells[1] if has_date else cells[0])
        interv = txt(cells[4] if has_date else cells[3])
        mh = mes_de(cells)
        # encabezado de mes: trae nombre de mes y NO es una fila de datos
        if mh and (not interv or not buseta or mes_de([buseta])):
            cur_y, cur_m = mh
            continue
        if es_encabezado_tabla(cells):
            continue
        if not buseta or not interv:
            continue
        # fecha
        fecha = ""
        if has_date and isinstance(cells[0], datetime.datetime):
            d = cells[0]
            fecha = "%04d-%02d-%02d" % (d.year, d.month, d.day)
        elif cur_m:
            fecha = "%04d-%02d-15" % (cur_y, cur_m)   # sin día → mitad de mes
        else:
            continue
        cc = txt(cells[2] if has_date else cells[1])
        km = txt(cells[3] if has_date else cells[2])
        seq[0] += 1
        records.append({
            "id": "seed|%d" % seq[0],
            "fecha": fecha,
            "placa": norm_buseta(buseta),     # el módulo agrupa "vehículos" por este campo
            "sistema": "",                    # los informes no traen subsistema
            "especifique": re.sub(r"\s+", " ", interv).strip(),  # descripción → clasificador
            "estado": "",                     # los informes no traen estado 1-10
            "cc": cc,
            "km": km,
            "fuente": "informe-historico",
        })
        added += 1
    return added, has_date


def main():
    files = sys.argv[1:] or DEFAULT_FILES
    records, seq = [], [0]
    for fn in files:
        if not os.path.exists(fn):
            print("  ⚠️  no existe:", fn); continue
        added, hd = parse_file(fn, records, seq)
        print("· %-55s %3d intervenciones (%s)" % (os.path.basename(fn), added, "con fecha" if hd else "por mes"))
    out = os.path.join(HERE, "taller_seed.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=1)
    # resumen por mes
    pormes = {}
    for r in records:
        pormes[r["fecha"][:7]] = pormes.get(r["fecha"][:7], 0) + 1
    print("\n✅ %d intervenciones → taller_seed.json" % len(records))
    for k in sorted(pormes):
        print("   %s : %d" % (k, pormes[k]))


if __name__ == "__main__":
    main()
