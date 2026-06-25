#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_vehiculos.py — Consolida la base de datos de vehículos TECC desde dos Excel:
  data_vehiculos/bd_veh_llantas.xlsx      (hoja 'BD' + 'REF LLANTAS')
  data_vehiculos/parque_automotor.xlsx    (hojas 'listado RC-RCE', 'TODO RIESGO',
                                            'Vencimientos T.O Zona Norte/Sur')
Salida: vehiculos.json  (clave = placa normalizada). Se embebe en la app vía build.py.
"""
import os, re, json, unicodedata, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
F1 = os.path.join(HERE, "data_vehiculos", "bd_veh_llantas.xlsx")
F2 = os.path.join(HERE, "data_vehiculos", "parque_automotor.xlsx")

def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).lower()

def placakey(s):
    if s is None: return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(s)).upper()

def clean(v):
    if v is None: return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    return "" if s.upper() in ("NULL", "NONE", "#N/A", "N/A", "N.A", "NA", "-", "--") else s

def read_sheet(path, sheet, header_row):
    """Devuelve (filas) como lista de dicts {header_normalizado: valor}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if header_row-1 >= len(rows): return []
    headers = [norm(h) for h in rows[header_row-1]]
    out = []
    for r in rows[header_row:]:
        d = {}
        for i, h in enumerate(headers):
            if not h: continue
            d[h] = clean(r[i]) if i < len(r) else ""
        out.append(d)
    return out

def pick(d, *names):
    for n in names:
        if n in d and d[n] != "": return d[n]
    return ""

VEH = {}  # placakey -> record

PLACA_RE = re.compile(r"^[A-Z]{3}\d{2,3}[A-Z]?$")   # placa colombiana (auto/moto)
def valid_placa(k):
    return bool(PLACA_RE.match(k))

def ensure(placa, numero="", create=True):
    k = placakey(placa)
    if not valid_placa(k): return None          # descarta chasis/VIN/filas basura (clave)
    if k not in VEH:
        if not create: return None
        VEH[k] = {"placa": placa.strip() if placa else "", "numero": "", "clase": "",
                  "marca": "", "linea": "", "modelo": "", "combustible": "", "capacidad": "",
                  "propietario": "", "propietarioReal": "", "empresaAfiliadora": "",
                  "empresaCliente": "", "ruta": "", "color": "", "conductor": "",
                  "refLlantas": "", "venc": {}, "fuentes": []}
    if numero and not VEH[k]["numero"]:
        VEH[k]["numero"] = str(numero).strip()
    return VEH[k]

def setif(rec, field, val):
    if val and not rec.get(field):
        rec[field] = val

# ---------- 1) Archivo 1, hoja 'BD' (base principal) ----------
for d in read_sheet(F1, "BD", 1):
    placa = pick(d, "placa")
    if not placakey(placa): continue
    rec = ensure(placa, pick(d, "n° interno", "no interno", "n interno", "interno"), create=True)
    if not rec: continue
    rec["fuentes"].append("BD")
    setif(rec, "clase", pick(d, "clase"))
    setif(rec, "marca", pick(d, "marca"))
    setif(rec, "modelo", pick(d, "modelo"))
    setif(rec, "propietario", pick(d, "propietario"))
    setif(rec, "capacidad", pick(d, "cap. pas.", "cap pas", "capacidad", "no. pasajeros"))
    setif(rec, "empresaAfiliadora", pick(d, "empresa afiliadora"))
    for label, col in [("soat","soat vto"),("rtm","r.t.m.a vto."),("to","t.o vto."),
                       ("bimensual","bimensual vto."),("extintor","extintor vto")]:
        v = pick(d, col)
        if v: rec["venc"][label] = v

# ---------- 1b) Archivo 1, hoja 'REF LLANTAS' ----------
try:
    wb = openpyxl.load_workbook(F1, read_only=True, data_only=True); ws = wb["REF LLANTAS"]
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or len(r) < 3: continue
        placa = r[2]
        rec = ensure(placa)
        if not rec: continue
        refs = [clean(c) for c in r[3:] if clean(c)]
        if refs: setif(rec, "refLlantas", " / ".join(dict.fromkeys(refs)))
    wb.close()
except Exception as e:
    print("REF LLANTAS:", e)

# ---------- 2) Archivo 2, hoja 'listado RC-RCE' ----------
for d in read_sheet(F2, "listado RC-RCE", 2):
    placa = pick(d, "placa")
    if not placakey(placa): continue
    rec = ensure(placa, pick(d, "# interno", "no. interno", "interno"), create=True)
    if not rec: continue
    rec["fuentes"].append("RC-RCE")
    setif(rec, "clase", pick(d, "clase"))
    setif(rec, "marca", pick(d, "marca"))
    setif(rec, "modelo", pick(d, "modelo"))
    setif(rec, "combustible", pick(d, "tipo combustible"))
    setif(rec, "capacidad", pick(d, "no. pasajeros", "no pasajeros"))
    setif(rec, "propietario", pick(d, "propietario"))

# ---------- 3) Archivo 2, hoja 'TODO RIESGO' ----------
for d in read_sheet(F2, "TODO RIESGO", 2):
    placa = pick(d, "placa")
    if not placakey(placa): continue
    rec = ensure(placa, pick(d, "nro_interno_afiliado", "# interno"))
    if not rec: continue
    rec["fuentes"].append("TODO-RIESGO")
    setif(rec, "clase", pick(d, "tipo_vehiculo_id"))
    setif(rec, "color", pick(d, "color"))
    setif(rec, "ruta", pick(d, "ruta"))
    setif(rec, "linea", pick(d, "linea"))
    setif(rec, "conductor", pick(d, "conductor"))
    setif(rec, "propietarioReal", pick(d, "propietario real"))
    setif(rec, "capacidad", pick(d, "capacidad"))

# ---------- 4) Vencimientos Zona Norte / Sur (empresa cliente) ----------
for sh in ["Vencimientos T.O Zona Norte", "Vencimientos T.O Zona Sur"]:
    try:
        for d in read_sheet(F2, sh, 2):
            placa = pick(d, "placa")
            if not placakey(placa): continue
            rec = ensure(placa, pick(d, "# interno", "no. interno"))
            if not rec: continue
            setif(rec, "empresaCliente", pick(d, "empresa cliente"))
            setif(rec, "conductor", pick(d, "conductores", "conductor"))
            setif(rec, "clase", pick(d, "clase"))
    except Exception as e:
        print(sh, e)

# ---------- normalización de clase ----------
CLASE_MAP = {
    "microbus": "Microbús", "micro bus": "Microbús", "micro": "Micro",
    "buseta": "Buseta", "bus": "Bus", "camioneta": "Camioneta",
    "automovil": "Automóvil", "auto": "Automóvil", "moto": "Moto",
    "campero": "Campero", "van": "Van", "personales": "Personales",
}
for rec in VEH.values():
    c = norm(rec["clase"])
    rec["clase"] = CLASE_MAP.get(c, "")   # whitelist: lo no reconocido (chasis/N/A) → sin clase
    rec["fuentes"] = sorted(set(rec["fuentes"]))

# ---------- quitar PII (el repo es PÚBLICO) ----------
PII = ("propietario", "propietarioReal", "conductor")
for r in VEH.values():
    for k in PII: r.pop(k, None)

# ---------- salida ----------
lst = sorted(VEH.values(), key=lambda r: (r["clase"] or "zz", r["numero"] or "zzz"))
out = {"generatedAt": "import_vehiculos", "count": len(lst), "vehiculos": lst}
with open(os.path.join(HERE, "vehiculos.json"), "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

# ---------- resumen ----------
from collections import Counter
print(f"✅ {len(lst)} vehículos consolidados → vehiculos.json")
clase_cnt = Counter(r["clase"] or "(sin clase)" for r in lst)
print("Clases:", dict(sorted(clase_cnt.items(), key=lambda x:-x[1])))
print("Con # interno:", sum(1 for r in lst if r["numero"]))
print("Con combustible:", sum(1 for r in lst if r["combustible"]),
      "| con ref llantas:", sum(1 for r in lst if r["refLlantas"]),
      "| con empresa cliente:", sum(1 for r in lst if r["empresaCliente"]))
